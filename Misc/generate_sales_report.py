import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import random
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Simulate a mock database with sales data
def create_mock_database():
    products = ['Laptop', 'Phone', 'Tablet', 'Headphones']
    dates = [(datetime(2025, 1, 1) + timedelta(days=x)).strftime('%Y-%m-%d') for x in range(30)]
    data = {
        'Date': [],
        'Product': [],
        'Sales': [],
        'Revenue': []
    }
    
    for date in dates:
        for product in products:
            data['Date'].append(date)
            data['Product'].append(product)
            sales = random.randint(10, 100)
            data['Sales'].append(sales)
            data['Revenue'].append(sales * random.uniform(50, 500))
    
    return pd.DataFrame(data)

# Pull data from mock database
def pull_data():
    return create_mock_database()

# Export data to Excel and embed charts
def export_to_excel(df, filename='sales_data_with_charts.xlsx'):
    # Save DataFrame to Excel
    df.to_excel(filename, index=False, sheet_name='Sales Data')
    
    # Create pie chart
    product_sales = df.groupby('Product')['Sales'].sum()
    plt.figure(figsize=(8, 6))
    plt.pie(product_sales, labels=product_sales.index, autopct='%1.1f%%', startangle=90)
    plt.title('Sales Distribution by Product')
    plt.axis('equal')
    pie_chart_path = 'product_sales_pie.png'
    plt.savefig(pie_chart_path)
    plt.close()
    
    # Create line graph
    plt.figure(figsize=(10, 6))
    for product in df['Product'].unique():
        product_data = df[df['Product'] == product]
        plt.plot(product_data['Date'], product_data['Sales'], label=product)
    plt.title('Sales Trend Over Time')
    plt.xlabel('Date')
    plt.ylabel('Sales')
    plt.legend()
    plt.xticks(rotation=45)
    plt.tight_layout()
    line_graph_path = 'sales_trend_line.png'
    plt.savefig(line_graph_path)
    plt.close()
    
    # Load workbook and add charts to a new sheet
    wb = load_workbook(filename)
    ws_charts = wb.create_sheet('Charts')
    
    # Add pie chart to Excel
    pie_img = Image(pie_chart_path)
    ws_charts.add_image(pie_img, 'A1')
    
    # Add line graph to Excel (positioned below pie chart)
    line_img = Image(line_graph_path)
    ws_charts.add_image(line_img, 'A20')
    
    wb.save(filename)
    print(f"Data and charts exported to {filename}")

# Main execution
def main():
    # Pull data
    df = pull_data()
    
    # Export to Excel with charts
    export_to_excel(df)

if __name__ == "__main__":
    main()